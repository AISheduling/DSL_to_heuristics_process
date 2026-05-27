import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path

from openai import OpenAI
from openpyxl import Workbook, load_workbook

api_key = os.environ.get("LITELLM_API_KEY")
if not api_key:
    raise ValueError("Не найден LITELLM_API_KEY! Установите переменную окружения.")

llm_client = OpenAI(
    api_key=api_key,
    base_url="https://api.duckduck.cloud/v1",
)

NUM_RUNS = 10

SYSTEM_PROMPT = """Ты эксперт-исследователь в области комбинаторных задач построения расписаний в условиях ограниченных ресурсов.

Тебе передаются метахарактеристики проекта в формате DSL описания. На их основе ты должен реализовать две Python-функции, которые будут вставлены в готовый каркас планировщика (Serial Schedule Generation Scheme).

* Функция 1: выбор следующей работы

```python
def select_job_fn(
    eligible_ids: list[int],
    jobs: dict[int, Job],
    snapshot: dict,
    meta: DSLMeta,
) -> int:
```

* Вызывается каждый раз, когда каркас формирует список доступных к запуску работ.
* Должна вернуть ровно один `job_id` из `eligible_ids`.
* Если вернуть id не из списка — каркас автоматически подставит `eligible_ids[0]`.

* Функция 2: выделение ресурсов

```python
def allocate_fn(
    job: Job,
    resources: dict[int, Resource],
    snapshot: dict,
    meta: DSLMeta,
) -> dict[int, int]:
```

* Вызывается сразу после выбора работы.
* Должна вернуть словарь `{resource_id: amount}` — сколько единиц каждого ресурса выделить.
* Выделяй ровно столько, сколько требует работа: `job.resources_required`.
* Если ресурсов не хватает — каркас сам вызовет `advance_time()`, повторять не нужно.

СТРУКТУРЫ ДАННЫХ, ДОСТУПНЫЕ В ФУНКЦИЯХ

### Job (dataclass)
```python
job.id: int
job.duration: int
job.predecessors: list[int]
job.successors: list[int]
job.resources_required: dict[int, int]
job.earliness_penalty: int
job.tardiness_penalty: int
job.non_renewable_consumption: dict[int, int]
job.selection_groups: list[list[int]]
```

### Resource (dataclass)
```python
resource.id: int
resource.name: str
resource.capacity: int
resource.r_type: str
resource.initial_stock: int
```

### snapshot (dict) — состояние планировщика в текущий момент
```python
snapshot["current_time"]: int
snapshot["completed_jobs"]: list[int]
snapshot["active_jobs"]: dict[int, int]
snapshot["eligible_jobs"]: list[int]
snapshot["renewable_used"]: dict[int, int]
snapshot["non_renewable_stock"]: dict[int, int]
snapshot["skipped_jobs"]: list[int]
```

### meta (DSLMeta) — метаданные задачи из DSL
```python
meta.project_id: str
meta.problem_type: str
meta.primary_metric: str
meta.primary_direction: str
meta.secondary_metric: str | None
meta.secondary_params: dict
meta.graph: dict
meta.has_non_renewable: bool
meta.is_multi_project: bool
meta.is_single_machine: bool
meta.conditional_scheduling: bool
meta.selection_groups_present: bool
```

Отвечай строго в JSON формате:
{
  "select_job_fn": "полный Python код функции select_job_fn",
  "allocate_fn": "полный Python код функции allocate_fn",
  "method": "название эвристики",
  "reasoning": "почему именно эта эвристика подходит для данной задачи, опираясь на метахарактеристики DSL (CPL, parallelism, order_strength и т.д.)",
  "description": "краткое техническое описание логики обеих функций"
}
"""


def load_dsl(filepath: str | Path) -> dict:
    with open(filepath, encoding="utf-8") as f:
        return json.load(f)


def generate_functions(dsl_content: dict, run_index: int) -> dict:
    """Отправляет DSL в LLM и возвращает сгенерированные функции."""
    user_message = (
        "На основе этого DSL-описания проекта реализуй функции "
        "select_job_fn и allocate_fn:\n\n"
        + json.dumps(dsl_content, ensure_ascii=False, indent=2)
    )

    response = llm_client.chat.completions.create(
        model="openai/gpt-5.4-nano",
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_message},
        ],
        temperature=0.7,
        max_tokens=5000,
        n=1,
        timeout=300,
        response_format={"type": "json_object"},
    )

    raw = response.choices[0].message.content
    parsed = json.loads(raw)

    return {
        "run_index": run_index,
        "tokens": {
            "prompt": response.usage.prompt_tokens,
            "completion": response.usage.completion_tokens,
            "total": response.usage.total_tokens,
        },
        "method": parsed.get("method"),
        "reasoning": parsed.get("reasoning"),
        "description": parsed.get("description"),
        "select_job_fn": parsed.get("select_job_fn"),
        "allocate_fn": parsed.get("allocate_fn"),
    }


def append_row_to_excel(xl_path: Path, row: list):
    if xl_path.exists():
        wb = load_workbook(xl_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append([
            "Run", "DSL файл", "Метод",
            "Prompt Tokens", "Completion Tokens", "Total Tokens",
            "Статус", "Описание",
        ])
    ws.append(row)
    wb.save(xl_path)


def run_experiment(dsl_filepath: Path, output_dir: Path, xl_path: Path,
                   num_runs: int = NUM_RUNS) -> list:
    dsl_content = load_dsl(dsl_filepath)
    dsl_name = dsl_filepath.stem
    results = []
    json_path = output_dir / f"{dsl_name}_results.json"

    for i in range(num_runs):
        print(f"Запуск {i + 1}/{num_runs}...")
        try:
            result = generate_functions(dsl_content, run_index=i + 1)
            results.append(result)
            print(
                f"Метод: {result['method']} | "
                f"Вход: {result['tokens']['prompt']} tok | "
                f"Выход: {result['tokens']['completion']} tok | "
                f"Всего: {result['tokens']['total']} tok"
            )
            row = [
                result["run_index"], dsl_name, result["method"],
                result["tokens"]["prompt"], result["tokens"]["completion"],
                result["tokens"]["total"], "OK", result["description"],
            ]
        except Exception as e:
            print(f"Ошибка на запуске {i + 1}: {e}")
            results.append({"run_index": i + 1, "error": str(e)})
            row = [i + 1, dsl_name, "N/A", "", "", "", "Ошибка", str(e)]

        append_row_to_excel(xl_path, row)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(
                {"dsl_file": dsl_filepath.name, "num_runs": num_runs, "runs": results},
                f, ensure_ascii=False, indent=2,
            )
        time.sleep(1)

    print(f"Готово. JSON сохранён: {json_path}")
    return results


if __name__ == "__main__":
    BASE_DIR = Path(__file__).resolve().parent.parent.parent

    # Путь к DSL: аргумент командной строки или все файлы из DSL_DIR
    if len(sys.argv) > 1:
        dsl_files = [Path(sys.argv[1])]
        print(f"Режим: одиночный файл -> {dsl_files[0]}")
    else:
        DSL_DIR = BASE_DIR / "data" / "references" / "brief DSL description of projects"
        dsl_files = sorted([
            f for f in DSL_DIR.glob("*.json")
            if not f.name.endswith("_results.json")
        ])
        print(f"Режим: все файлы из {DSL_DIR}")

    if not dsl_files:
        print("JSON-файлы не найдены.")
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    OUTPUT_DIR = BASE_DIR / "experiments" / "experiment_skeleton" / timestamp
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    XL_DIR = OUTPUT_DIR / "experiment_results.xlsx"

    print(f"Найдено файлов: {len(dsl_files)}")
    print(f"Результаты будут сохранены в: {OUTPUT_DIR}\n")

    all_results = {}
    for dsl_file in dsl_files:
        print(f"{'-' * 55}")
        print(f"Обработка: {dsl_file.name}")
        all_results[dsl_file.name] = run_experiment(
            dsl_filepath=dsl_file,
            output_dir=OUTPUT_DIR,
            xl_path=XL_DIR,
        )

    summary_path = OUTPUT_DIR / "all_results.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)

    print(f"\nОбщие результаты: {summary_path}")
    print(f"Excel таблица: {XL_DIR}")
    print(f"\n{'-' * 55}")
    print("Итого по методам:")
    for filename, runs in all_results.items():
        rules = [r.get("method", "N/A") for r in runs if "error" not in r]
        print(f"  {filename}: {rules}")